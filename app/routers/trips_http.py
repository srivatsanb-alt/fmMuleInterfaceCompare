import logging
from typing import Union
from fastapi import APIRouter, Depends, Request
from fastapi.encoders import jsonable_encoder
from fastapi.responses import StreamingResponse
from sqlalchemy.orm.attributes import flag_modified
import asyncio
import pandas as pd
from datetime import datetime, time, date
import io
import math
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ati code imports
import app.routers.dependencies as dpd
import models.request_models as rqm
import models.trip_models as tm
import models.misc_models as mm
from models.db_session import DBSession
from utils.util import str_to_dt, dt_to_str
import utils.trip_utils as tu
import core.constants as cc
import utils.util as utils_util
import core.common as ccm
from openpyxl import Workbook
from openpyxl.worksheet.page import PageMargins
from utils.auth_utils import AuthValidator


router = APIRouter(
    prefix="/api/v1/trips",
    tags=["trips"],
    responses={404: {"description": "Not found"}},
)

# to book, delete ongoing and pending trips, clear optimal dispatch assignments,
# posts trip status on the FM and also does trip analysis.

@router.post("/book")
async def book(request: Request, booking_req: rqm.BookingReq, user=Depends(AuthValidator('manage_trip'))):
    user_name = user["user_name"]
    response = await dpd.process_req_with_response(None, booking_req, user_name)
    return response


@router.delete("/force_delete/ongoing/{sherpa_name}")
async def force_delete_ongoing_trip(
    sherpa_name: str, user=Depends(AuthValidator('manage_trip'))
):
    response = {}

    if not user:
        dpd.raise_error("Unknown requester", 401)
    user_name = user["user_name"]

    with DBSession(engine=ccm.engine) as dbsession:
        sherpa = dbsession.get_sherpa(sherpa_name)

        if sherpa.status.disabled_reason not in [
            cc.DisabledReason.STALE_HEARTBEAT,
            cc.DisabledReason.SOFTWARE_NOT_COMPATIBLE,
        ]:
            dpd.raise_error("This option can be used only if the sherpa is disconnected")

        if sherpa.status.trip_id is None:
            dpd.raise_error(f"{sherpa_name} has no ongoing trip")

    force_delete_ongoing_trip_req = rqm.ForceDeleteOngoingTripReq(sherpa_name=sherpa_name)

    response = await dpd.process_req_with_response(
        None, force_delete_ongoing_trip_req, user_name
    )

    return response


@router.delete("/ongoing/{booking_id}/{trip_id}")
async def delete_ongoing_trip(
    booking_id: int, trip_id: int, user=Depends(AuthValidator('manage_trip'))
):
    if not user:
        dpd.raise_error("Unknown requester", 401)
    user_name = user["user_name"]

    with DBSession(engine=ccm.engine) as dbsession:
        trips = dbsession.get_trip_with_booking_id(booking_id)
        if not trips:
            dpd.raise_error("no trip with the given booking_id")

        if trip_id == -1:
            delete_ongoing_trip_req: rqm.DeleteOngoingTripReq = rqm.DeleteOngoingTripReq(
                booking_id=booking_id
            )
        else:
            valid_trip_id = False
            for trip in trips:
                if trip_id == trip.id:
                    valid_trip_id = True

            if not valid_trip_id:
                dpd.raise_error(
                    f"Invalid detail, no trip (trip_id: {trip_id}) for booking_id: {booking_id}"
                )
            delete_ongoing_trip_req: rqm.DeleteOngoingTripReq = rqm.DeleteOngoingTripReq(
                booking_id=booking_id, trip_id=trip_id
            )

    response = await dpd.process_req_with_response(None, delete_ongoing_trip_req, user_name)

    return response


@router.delete("/booking/{booking_id}/{trip_id}")
async def delete_pending_trip(
    booking_id: int, trip_id: int, user=Depends(AuthValidator('manage_trip'))
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)
    user_name = user["user_name"]

    with DBSession(engine=ccm.engine) as dbsession:
        trips = dbsession.get_trip_with_booking_id(booking_id)
        paused_trip = dbsession.get_paused_trip_with_booking_id(booking_id)

        if paused_trip:
            dbsession.session.delete(paused_trip)
            return response
        
        if not trips:
            dpd.raise_error("no trip with the given booking_id")

        if trip_id == -1:
            delete_booked_trip_req: rqm.DeleteBookedTripReq = rqm.DeleteBookedTripReq(
                booking_id=booking_id
            )

        else:
            valid_trip_id = False
            for trip in trips:
                if trip_id == trip.id:
                    valid_trip_id = True

            if not valid_trip_id:
                dpd.raise_error(
                    f"Invalid detail, no pending trip (trip_id: {trip_id}) for booking_id: {booking_id}"
                )

            delete_booked_trip_req: rqm.DeleteBookedTripReq = rqm.DeleteBookedTripReq(
                booking_id=booking_id, trip_id=trip_id
            )

    response = await dpd.process_req_with_response(None, delete_booked_trip_req, user_name)

    return response


# based on the trip bookings, optimal dispatch assigns the tasks to various sherpas optimally.
# deletes all the assignments done to the sherpas.


@router.get("/booking/{entity_name}/clear_optimal_dispatch_assignments")
async def clear_optimal_dispatch_assignments(
    entity_name=Union[str, None], user=Depends(AuthValidator('fm'))
):
    response = {}

    if not user:
        dpd.raise_error("Unknown requester", 401)

    if not entity_name:
        dpd.raise_error("No entity name")
    
    user_name = user["user_name"]

    delete_optimal_dispatch_assignments_req = rqm.DeleteOptimalDispatchAssignments(
        fleet_name=entity_name
    )

    response = await dpd.process_req_with_response(
        None, delete_optimal_dispatch_assignments_req, user_name
    )

    return response


@router.post("/status/{type}")
async def trip_status_with_type(
    type: str,
    trip_status_req: rqm.TripStatusReq,
    user=Depends(AuthValidator('fm')),
):
    response = {}

    if not user:
        dpd.raise_error("Unknown requester", 401)

    valid_status = []
    if type == "yet_to_start":
        valid_status = tm.YET_TO_START_TRIP_STATUS
    elif type == "completed":
        valid_status = tm.COMPLETED_TRIP_STATUS
    elif type == "ongoing":
        valid_status = tm.ONGOING_TRIP_STATUS
    else:
        dpd.raise_error("Query sent for an invalid trip type")

    with DBSession(engine=ccm.engine) as dbsession:
        if trip_status_req.from_dt and trip_status_req.to_dt:
            trip_status_req.from_dt = str_to_dt(trip_status_req.from_dt)
            trip_status_req.to_dt = str_to_dt(trip_status_req.to_dt)

            all_trips = dbsession.get_trips_with_timestamp_and_status(
                trip_status_req.from_dt, trip_status_req.to_dt, valid_status
            )

        else:
            if not trip_status_req.trip_ids:
                return response

            all_trips = dbsession.get_trips_with_ids_and_status(
                trip_status_req.trip_ids, valid_status
            )
            # dpd.raise_error("no trip id given or available in the given timeframe")

        count = 0
        for trip in all_trips:
            response.update({trip.id: tu.get_trip_status(trip)})

            # introducing sleep to allow other endpoints to work simultaneously
            count = count + 1
            if count % 50 == 0:
                await asyncio.sleep(100e-3)

    return response


@router.post("/status_pg/{type}")
async def trip_status_pg_with_type(
    type: str,
    trip_status_req: rqm.TripStatusReq_pg,
    user=Depends(AuthValidator('manage_trip')),
):
    response = {}

    if not user:
        dpd.raise_error("Unknown requester", 401)

    valid_status = []
    if type == "yet_to_start" or type == "scheduled":
        valid_status = tm.YET_TO_START_TRIP_STATUS
    elif type == "completed":
        valid_status = tm.COMPLETED_TRIP_STATUS
    elif type == "ongoing":
        valid_status = tm.ONGOING_TRIP_STATUS
    elif type == "active":
        valid_status = tm.ACTIVE_TRIP_STATUS
    else:
        dpd.raise_error("Query sent for an invalid trip type")

    logging.getLogger("uvicorn").info(
        f"trip_status_req: {jsonable_encoder(trip_status_req)}"
    )

    with DBSession(engine=ccm.engine) as dbsession:
        if trip_status_req.from_dt and trip_status_req.to_dt:
            trip_status_req.from_dt = str_to_dt(trip_status_req.from_dt)
            trip_status_req.to_dt = str_to_dt(trip_status_req.to_dt)

        if type == "scheduled":
            response = dbsession.get_scheduled_trips(
                trip_status_req.filter_fleets,
                valid_status,
                trip_status_req.filter_status,
                trip_status_req.search_txt,
                trip_status_req.sort_field,
                trip_status_req.sort_order,
                trip_status_req.page_no,
                trip_status_req.rec_limit,
            )
        else:
            response = dbsession.get_trips_with_timestamp_and_status_pagination(
                trip_status_req.from_dt,
                trip_status_req.to_dt,
                trip_status_req.filter_fleets,
                valid_status,
                trip_status_req.filter_sherpa_names,
                trip_status_req.filter_status,
                trip_status_req.booked_by,
                trip_status_req.search_by_stations,
                trip_status_req.search_txt,
                trip_status_req.sort_field,
                trip_status_req.sort_order,
                trip_status_req.page_no,
                trip_status_req.rec_limit,
            )

    return response


@router.post("/status")
async def trip_status(
    trip_status_req: rqm.TripStatusReq, user=Depends(AuthValidator('fm'))
):
    response = {}

    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        all_trips = None
        if trip_status_req.from_dt and trip_status_req.to_dt:
            trip_status_req.from_dt = str_to_dt(trip_status_req.from_dt)
            trip_status_req.to_dt = str_to_dt(trip_status_req.to_dt)
            all_trips = dbsession.get_trips_with_timestamp(
                trip_status_req.from_dt, trip_status_req.to_dt
            )

        else:
            if not trip_status_req.trip_ids:
                return response
            all_trips = dbsession.get_trips_with_ids(trip_status_req.trip_ids)

        count = 0
        for trip in all_trips:
            response.update({trip.id: tu.get_trip_status(trip)})
            count = count + 1

            # introducing sleep to allow other endpoints to work simultaneously
            if count % 50 == 0:
                await asyncio.sleep(100e-3)

    return response


# returns ongoing trip status.


@router.get("/ongoing_trip_status")
async def ongoing_trip_status(user=Depends(AuthValidator('fm'))):
    if not user:
        dpd.raise_error("Unknown requester", 401)

    response = {}
    with DBSession(engine=ccm.engine) as dbsession:
        all_ongoing_trips = dbsession.get_all_ongoing_trips()
        for ongoing_trip in all_ongoing_trips:
            response.update({ongoing_trip.trip_id: tu.get_trip_status(ongoing_trip.trip)})

    return response


# returns the complete analysis of the trip as in when it started, when it ended, when was it supposed to end
# performs analysis for every trip leg and also for the overall trip.


@router.post("/analytics_pg")
async def trip_analytics_pg(
    trip_analytics_req: rqm.TripStatusReq_pg, user_name=Depends(AuthValidator('analytics'))
):
    response = {}
    if not user_name:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        if trip_analytics_req.from_dt and trip_analytics_req.to_dt:
            trip_analytics_req.from_dt = str_to_dt(trip_analytics_req.from_dt)
            trip_analytics_req.to_dt = str_to_dt(trip_analytics_req.to_dt)

        trip_analytics = dbsession.get_trip_analytics_with_pagination(
            trip_analytics_req.from_dt,
            trip_analytics_req.to_dt,
            trip_analytics_req.filter_fleets,
            trip_analytics_req.filter_sherpa_names,
            trip_analytics_req.filter_status,
            trip_analytics_req.sort_field,
            trip_analytics_req.sort_order,
            trip_analytics_req.page_no,
            trip_analytics_req.rec_limit,
        )
        response = trip_analytics

    return response


@router.post("/analytics")
async def trip_analytics(
    trip_analytics_req: rqm.TripStatusReq, user=Depends(AuthValidator('analytics'))
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        if trip_analytics_req.from_dt and trip_analytics_req.to_dt:
            trip_analytics_req.from_dt = str_to_dt(trip_analytics_req.from_dt)
            trip_analytics_req.to_dt = str_to_dt(trip_analytics_req.to_dt)

            all_trip_analytics = dbsession.get_trip_analytics_with_timestamp(
                trip_analytics_req.from_dt, trip_analytics_req.to_dt
            )

        else:
            if not trip_analytics_req.trip_ids:
                return response
                all_trip_analytics = dbsession.get_trip_analytics_with_trip_ids(
                    trip_analytics_req.trip_ids
                )
            # dpd.raise_error("no trip id given or available in the given timeframe")

        count = 0
        for trip_analytics in all_trip_analytics:
            response.update(
                {trip_analytics.trip_leg_id: tu.get_trip_analytics(trip_analytics)}
            )

            # introducing sleep to allow other endpoints to work simultaneously
            count = count + 1
            if count % 50 == 0:
                await asyncio.sleep(100e-3)

    return response


@router.post("/analytics_with_trip_info")
async def trip_analytics_with_trip_info(
    trip_analytics_req: rqm.TripAnalyticsWithTripInfoReq, user_name=Depends(AuthValidator('fm'))
):
    response = {}
    if not user_name:
        dpd.raise_error("Unknown requester", 401)    

    with DBSession(engine=ccm.engine) as dbsession:
        all_trip_analytics = dbsession.get_trip_analytics_with_trips_info(
            trip_analytics_req.from_dt, trip_analytics_req.to_dt, trip_analytics_req.booked_by, trip_analytics_req.filter_fleets
        )
        response = all_trip_analytics

    return response

@router.post("/{trip_id}/add_trip_metadata")
async def add_trip_metadata(
    trip_id: int,
    new_trip_metadata: rqm.TripMetaData,
    user=Depends(AuthValidator('manage_trip')),
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        trip: tm.Trip = dbsession.get_trip(trip_id)
        if trip.trip_metadata is None:
            trip.trip_metadata = {}

        trip.trip_metadata.update(new_trip_metadata.metadata)
        flag_modified(trip, "trip_metadata")

    return response


@router.get("/{trip_id}/add_trip_description/{description}")
async def add_trip_description(
    trip_id: int, description: str, user=Depends(AuthValidator('manage_trip'))
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        trip: tm.Trip = dbsession.get_trip(trip_id)
        if trip.trip_metadata is None:
            trip.trip_metadata = {}
        trip.trip_metadata.update({"description": description})
        flag_modified(trip, "trip_metadata")

    return response


@router.get("/popular_route/{fleet_name}/{num_routes}")
async def populate_route(
    fleet_name: str,
    num_routes: int,
    user=Depends(AuthValidator('manage_trip')),
):
    response = []

    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        populate_routes = dbsession.get_popular_routes(fleet_name)

    for route in populate_routes[:num_routes]:
        response.extend(route)

    return response


@router.post("/save_route")
async def save_route(
    save_route_req: rqm.SaveRouteReq, user=Depends(AuthValidator('manage_trip'))
):
    if not user:
        dpd.raise_error("Unknown requester", 401)
    user_name = user["user_name"]

    response = await dpd.process_req_with_response(None, save_route_req, user_name)

    return response


@router.get("/get_saved_routes/{fleet_name}/{backend_usage}")
async def get_saved_routes(
    fleet_name: str, backend_usage: bool, user=Depends(AuthValidator('manage_trip'))
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        saved_routes = dbsession.get_saved_routes_fleet(fleet_name)

        for saved_route in saved_routes:
            used_by_backend = False
            update = False

            for _tag in mm.ConditionalTripTags:
                if _tag in saved_route.tag:
                    used_by_backend = True

            if used_by_backend and backend_usage:
                update = True
            elif not used_by_backend and not backend_usage:
                update = True

            if update:
                response.update(
                    {
                        saved_route.tag: {
                            "route": saved_route.route,
                            "fleet_name": saved_route.fleet_name,
                            "other_info": saved_route.other_info,
                        }
                    }
                )

    return response


@router.get("/get_saved_route/{route_tag}")
async def get_saved_route(route_tag: str, user=Depends(AuthValidator('manage_trip'))):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        saved_route = dbsession.get_saved_route(route_tag)

        if saved_route is None:
            dpd.raise_error(f"route with tag {route_tag} does not exist")

        response = utils_util.get_table_as_dict(tm.SavedRoutes, saved_route)

    return response


@router.delete("/delete_saved_route/{tag}")
async def delete_saved_route(tag: str, user=Depends(AuthValidator('manage_trip'))):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        saved_route = dbsession.get_saved_route(tag)

        if saved_route is None:
            dpd.raise_error(f"No saved route with tag:{tag}")

        can_edit = saved_route.other_info.get("can_edit", "False")

        if not eval(can_edit):
            dpd.raise_error("Cannot edit/delete this route tag, can_edit is set to False")

        dbsession.session.delete(saved_route)

    return response


@router.post("/update_saved_route_info")
async def update_saved_route_metadata(
    update_saved_route_req: rqm.UpdateSavedRouteReq,
    user=Depends(AuthValidator('manage_trip')),
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        saved_route = dbsession.get_saved_route(update_saved_route_req.tag)

        if saved_route is None:
            dpd.raise_error(f"No saved route with tag:{update_saved_route_req.tag}")

        if saved_route.other_info is None:
            saved_route.other_info = {}

        saved_route.other_info.update(update_saved_route_req.other_info)

        flag_modified(saved_route, "other_info")

    return response

def format_excel_sheet(worksheet, df):
    """Apply formatting to Excel worksheet"""
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Only format if there are rows in the worksheet
    if worksheet.max_row > 0:
        # Format header row
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment

        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            # Set minimum width of 10 and maximum of 50
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

        # Apply borders to all data cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
                if cell.row > 1:  # Data rows (not header)
                    cell.alignment = Alignment(vertical='center')

        # Freeze the header row (only if there are multiple rows)
        if worksheet.max_row > 1:
            worksheet.freeze_panes = 'A2'

@router.post("/export_all_analytics_data")
async def export_all_analytics_data(
    trip_analytics_req: rqm.TripAnalyticsReq,
    user=Depends(dpd.get_user_from_header)
):
    if not user:
        dpd.raise_error("Unknown requester", 401)
    response = {}
    with DBSession() as dbsession:
        if trip_analytics_req.start_time and trip_analytics_req.end_time:
            # Convert string dates to datetime objects
            trip_analytics_req.start_time = utils_util.str_to_dt(trip_analytics_req.start_time)
            trip_analytics_req.end_time = utils_util.str_to_dt(trip_analytics_req.end_time)
            all_trip_analytics = dbsession.get_trip_analytics_to_export(
                trip_analytics_req.start_time,
                trip_analytics_req.end_time,
                trip_analytics_req.fleet_name,
                trip_analytics_req.status,
                trip_analytics_req.sherpa_name,
                trip_analytics_req.sort_field,
                trip_analytics_req.sort_order,
            )
        else:
            dpd.raise_error("start_time and end_time are required")
        
        # Prepare data for two separate sheets
        trips_data = []
        trip_legs_data = []
        
        for trip_analytic in all_trip_analytics:
            trip_analytic_legs = trip_analytic.get("legs", [])
            
            # Create trip data (without legs)
            trip_data = {**trip_analytic}
            del trip_data["legs"]
            
            # Split datetime fields into date and time before formatting
            from datetime import datetime
            
            # Booking time - split into date and time
            if trip_data.get('booking_time'):
                try:
                    if isinstance(trip_data['booking_time'], str):
                        booking_dt = datetime.fromisoformat(trip_data['booking_time'].replace('Z', '+00:00'))
                    else:
                        booking_dt = trip_data['booking_time']
                    trip_data['booking_time_only'] = booking_dt.strftime('%H:%M:%S')  # Time only
                    # Keep original for date extraction after format_dates
                except:
                    trip_data['booking_time_only'] = ''
            else:
                trip_data['booking_time_only'] = ''
            
            # Start time - split into date and time
            if trip_data.get('start_time'):
                try:
                    if isinstance(trip_data['start_time'], str):
                        start_dt = datetime.fromisoformat(trip_data['start_time'].replace('Z', '+00:00'))
                    else:
                        start_dt = trip_data['start_time']
                    trip_data['start_time_only'] = start_dt.strftime('%H:%M:%S')  # Time only
                    # Keep original for date extraction after format_dates
                except:
                    trip_data['start_time_only'] = ''
            else:
                trip_data['start_time_only'] = ''
            
            # End time - split into date and time
            if trip_data.get('end_time'):
                try:
                    if isinstance(trip_data['end_time'], str):
                        end_dt = datetime.fromisoformat(trip_data['end_time'].replace('Z', '+00:00'))
                    else:
                        end_dt = trip_data['end_time']
                    trip_data['end_time_only'] = end_dt.strftime('%H:%M:%S')  # Time only
                    # Keep original for date extraction after format_dates
                except:
                    trip_data['end_time_only'] = ''
            else:
                trip_data['end_time_only'] = ''
            
            # Now format the dates (this will convert datetime objects to strings)
            trip_data = utils_util.format_dates(trip_data)
            
            # Extract date parts from the formatted datetime strings
            # Booking time
            if trip_data.get('booking_time'):
                try:
                    # Parse the formatted date string and extract just the date part
                    if ' ' in trip_data['booking_time']:
                        trip_data['booking_time'] = trip_data['booking_time'].split(' ')[0]
                except:
                    pass
            
            # Start time
            if trip_data.get('start_time'):
                try:
                    # Parse the formatted date string and extract just the date part
                    if ' ' in trip_data['start_time']:
                        trip_data['start_time'] = trip_data['start_time'].split(' ')[0]
                except:
                    pass
            
            # End time
            if trip_data.get('end_time'):
                try:
                    # Parse the formatted date string and extract just the date part
                    if ' ' in trip_data['end_time']:
                        trip_data['end_time'] = trip_data['end_time'].split(' ')[0]
                except:
                    pass
            
            # Calculate additional fields from trip legs
            # Expected Trip Time is sum of etas_at_start array
            etas_at_start = trip_data.get('etas_at_start', [])
            expected_trip_time = sum(etas_at_start) if etas_at_start else 0
            
            total_trip_time = 0
            total_stoppage_time = 0
            total_obstacle_stoppage_time = 0
            total_visa_stoppage_time = 0
            total_error_wait_time = 0
            avg_progress = 0
            
            if trip_analytic_legs:
                # Calculate aggregated values from trip legs
                for leg in trip_analytic_legs:
                    total_trip_time += leg.get('actual_trip_time', 0) or 0
                    total_stoppage_time += (leg.get('time_elapsed_other_stoppages', 0) or 0) + \
                                         (leg.get('time_elapsed_obstacle_stoppages', 0) or 0) + \
                                         (leg.get('time_elapsed_visa_stoppages', 0) or 0)
                    total_obstacle_stoppage_time += leg.get('time_elapsed_obstacle_stoppages', 0) or 0
                    total_visa_stoppage_time += leg.get('time_elapsed_visa_stoppages', 0) or 0
                
                # Calculate average progress (convert from decimal to percentage)
                progress_values = [leg.get('progress', 0) or 0 for leg in trip_analytic_legs if leg.get('progress') is not None]
                avg_progress = (sum(progress_values) / len(progress_values) * 100) if progress_values else 0
                
                # Calculate error wait time from sherpa_mode_change table
                if trip_data.get('start_time') and trip_data.get('end_time') and trip_data.get('sherpa_name'):
                    try:
                        from models import misc_models as mm
                        from sqlalchemy import and_
                        
                        error_mode_changes = dbsession.session.query(mm.SherpaModeChange).filter(
                            and_(
                                mm.SherpaModeChange.sherpa_name == trip_data['sherpa_name'],
                                mm.SherpaModeChange.mode == 'error',
                                mm.SherpaModeChange.started_at >= trip_data['start_time'],
                                mm.SherpaModeChange.started_at <= trip_data['end_time']
                            )
                        ).all()
                        
                        for error_change in error_mode_changes:
                            if error_change.ended_at:
                                # Error period ended during trip
                                error_duration = (error_change.ended_at - error_change.started_at).total_seconds()
                            else:
                                # Error period started during trip but didn't end
                                error_duration = (trip_data['end_time'] - error_change.started_at).total_seconds()
                            
                            total_error_wait_time += max(0, error_duration)
                    except Exception as e:
                        # If there's an error calculating error wait time, continue with 0
                        total_error_wait_time = 0
            
            # Add calculated fields to trip data
            trip_data['expected_trip_time_calculated'] = expected_trip_time
            trip_data['total_trip_time_calculated'] = total_trip_time
            trip_data['total_obstacle_stoppage_time_calculated'] = total_obstacle_stoppage_time
            trip_data['total_visa_stoppage_time_calculated'] = total_visa_stoppage_time
            trip_data['total_error_wait_time_calculated'] = total_error_wait_time
            trip_data['progress_calculated'] = avg_progress
            
            # Extract data from trip_metadata if available
            trip_metadata = trip_data.get('trip_metadata', {})
            if isinstance(trip_metadata, dict):
                # Get progress from trip_metadata if available, otherwise use calculated average
                metadata_progress = trip_metadata.get('total_trip_progress')
                if metadata_progress is not None:
                    try:
                        # Convert to percentage if it's a string or number
                        progress_value = float(metadata_progress)
                        trip_data['progress_calculated'] = progress_value
                    except (ValueError, TypeError):
                        pass  # Keep the calculated average if conversion fails
                
                # Get total dispatch wait time from trip_metadata
                dispatch_wait_time = trip_metadata.get('total_dispatch_wait_time')
                if dispatch_wait_time is not None:
                    try:
                        trip_data['total_dispatch_wait_time_calculated'] = float(dispatch_wait_time)
                    except (ValueError, TypeError):
                        trip_data['total_dispatch_wait_time_calculated'] = 0
                else:
                    trip_data['total_dispatch_wait_time_calculated'] = 0
            else:
                trip_data['total_dispatch_wait_time_calculated'] = 0
            
            # Calculate total stoppage time as sum of all stoppage types
            trip_data['total_stoppage_time_calculated'] = (
                trip_data['total_obstacle_stoppage_time_calculated'] +
                trip_data['total_visa_stoppage_time_calculated'] +
                trip_data['total_dispatch_wait_time_calculated'] +
                trip_data['total_error_wait_time_calculated']
            )
            
            # Convert scheduled boolean to Yes/No
            if trip_data.get('scheduled') is True:
                trip_data['scheduled'] = 'Yes'
            elif trip_data.get('scheduled') is False:
                trip_data['scheduled'] = 'No'
            
            
            # Convert all list and dict fields to strings for Excel compatibility
            for key, value in trip_data.items():
                if isinstance(value, list):
                    if key == 'route':
                        # Special formatting for route: use arrows instead of commas
                        trip_data[key] = ' → '.join(map(str, value))
                    elif key == 'route_lengths':
                        # Sum the route lengths instead of showing individual values
                        try:
                            trip_data[key] = sum(float(x) for x in value if x is not None)
                        except (ValueError, TypeError):
                            trip_data[key] = 0
                    else:
                        trip_data[key] = ', '.join(map(str, value))
                elif isinstance(value, dict):
                    trip_data[key] = str(value)
            
            trips_data.append(trip_data)
            
            # Create trip legs data with only the fields we want
            for trip_analytic_leg in trip_analytic_legs:
                # Create a new leg_data dict with only the fields we want
                from_station = trip_analytic_leg.get('from_station', '')
                to_station = trip_analytic_leg.get('to_station', '')
                
                # Create trip leg description with proper handling of empty from_station
                if from_station and to_station:
                    trip_leg = f"{from_station} → {to_station}"
                elif to_station:
                    trip_leg = f"→ {to_station}"  # Show arrow + to_station if from_station is empty
                elif from_station:
                    trip_leg = from_station  # Show only from_station if to_station is empty
                else:
                    trip_leg = ''  # Both are empty
                
                # Get dispatch wait time from trip metadata
                trip_metadata = trip_analytic.get('trip_metadata', {})
                total_dispatch_wait_time = 0
                if isinstance(trip_metadata, dict):
                    dispatch_wait = trip_metadata.get('total_dispatch_wait_time')
                    if dispatch_wait is not None:
                        try:
                            total_dispatch_wait_time = float(dispatch_wait)
                        except (ValueError, TypeError):
                            total_dispatch_wait_time = 0
                
                # Calculate dispatch wait time per leg (distribute evenly across all legs)
                num_legs = len(trip_analytic_legs)
                leg_dispatch_wait_time = total_dispatch_wait_time / num_legs if num_legs > 0 else 0
                
                leg_data = {
                    'fleet_name': trip_analytic.get('fleet_name'),
                    'trip_id': trip_analytic_leg.get('trip_id'),
                    'booking_id': trip_analytic.get('booking_id'),
                    'trip_leg_id': trip_analytic_leg.get('trip_leg_id'),
                    'trip_leg': trip_leg,
                    'start_time': trip_analytic_leg.get('start_time'),
                    'end_time': trip_analytic_leg.get('end_time'),
                    'actual_trip_time': trip_analytic_leg.get('actual_trip_time'),
                    'time_elapsed_obstacle_stoppages': trip_analytic_leg.get('time_elapsed_obstacle_stoppages'),
                    'time_elapsed_visa_stoppages': trip_analytic_leg.get('time_elapsed_visa_stoppages'),
                    'leg_dispatch_wait_time': leg_dispatch_wait_time
                }
                
                # Split datetime fields into date and time before formatting
                from datetime import datetime
                
                # Start time - split into date and time
                if leg_data.get('start_time'):
                    try:
                        if isinstance(leg_data['start_time'], str):
                            start_dt = datetime.fromisoformat(leg_data['start_time'].replace('Z', '+00:00'))
                        else:
                            start_dt = leg_data['start_time']
                        leg_data['start_time_only'] = start_dt.strftime('%H:%M:%S')  # Time only
                        # Keep original for date extraction after format_dates
                    except:
                        leg_data['start_time_only'] = ''
                else:
                    leg_data['start_time_only'] = ''
                
                # End time - split into date and time
                if leg_data.get('end_time'):
                    try:
                        if isinstance(leg_data['end_time'], str):
                            end_dt = datetime.fromisoformat(leg_data['end_time'].replace('Z', '+00:00'))
                        else:
                            end_dt = leg_data['end_time']
                        leg_data['end_time_only'] = end_dt.strftime('%H:%M:%S')  # Time only
                        # Keep original for date extraction after format_dates
                    except:
                        leg_data['end_time_only'] = ''
                else:
                    leg_data['end_time_only'] = ''
                
                # Format dates
                leg_data = utils_util.format_dates(leg_data)
                
                # Extract date parts from the formatted datetime strings
                # Start time
                if leg_data.get('start_time'):
                    try:
                        # Parse the formatted date string and extract just the date part
                        if ' ' in leg_data['start_time']:
                            leg_data['start_time'] = leg_data['start_time'].split(' ')[0]
                    except:
                        pass
                
                # End time
                if leg_data.get('end_time'):
                    try:
                        # Parse the formatted date string and extract just the date part
                        if ' ' in leg_data['end_time']:
                            leg_data['end_time'] = leg_data['end_time'].split(' ')[0]
                    except:
                        pass
                
                # Convert any list and dict fields to strings for Excel compatibility
                for key, value in leg_data.items():
                    if isinstance(value, list):
                        leg_data[key] = ', '.join(map(str, value))
                    elif isinstance(value, dict):
                        leg_data[key] = str(value)
                
                trip_legs_data.append(leg_data)
        
        # Create Excel workbook with two sheets
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define expected columns for trips in the desired order
        expected_trip_columns = [
            'fleet_name', 'id', 'booking_id', 'route', 'route_lengths', 'sherpa_name',
            'booking_time', 'booking_time_only', 'start_time', 'start_time_only', 
            'end_time', 'end_time_only', 'status', 'scheduled',
            'expected_trip_time_calculated', 'total_trip_time_calculated', 'progress_calculated',
            'total_stoppage_time_calculated', 'total_obstacle_stoppage_time_calculated',
            'total_dispatch_wait_time_calculated', 'total_visa_stoppage_time_calculated',
            'total_error_wait_time_calculated', 'booked_by'
        ]
        
        # Create Trips sheet
        trips_df = pd.DataFrame(trips_data)
        
        # Reorder DataFrame columns to match the expected order
        # Only include columns that exist in the data
        available_columns = [col for col in expected_trip_columns if col in trips_df.columns]
        trips_df = trips_df[available_columns]
        
        trips_ws = wb.create_sheet("Trips")
        
        # Always ensure headers are present, even with empty data
        if trips_df.empty:
            # Create empty DataFrame with expected columns for trips
            trips_df = pd.DataFrame(columns=expected_trip_columns)
        
        # Define user-friendly column headers
        column_headers = {
            'fleet_name': 'Fleet',
            'id': 'Trip ID',
            'booking_id': 'Booking ID',
            'route': 'Route',
            'route_lengths': 'Route Length (m)',
            'sherpa_name': 'Sherpa',
            'booking_time': 'Trip Booking Date',
            'booking_time_only': 'Trip Booking Time',
            'start_time': 'Trip Start Date',
            'start_time_only': 'Trip Start Time',
            'end_time': 'Trip End Date',
            'end_time_only': 'Trip End Time',
            'status': 'Status',
            'scheduled': 'Schedule',
            'expected_trip_time_calculated': 'Expected Trip Time (s)',
            'total_trip_time_calculated': 'Trip Time (s)',
            'progress_calculated': 'Progress (%)',
            'total_stoppage_time_calculated': 'Total Stoppage Time (s)',
            'total_obstacle_stoppage_time_calculated': 'Total Obstacle Stoppage Time (s)',
            'total_dispatch_wait_time_calculated': 'Total Dispatch Wait Time (s)',
            'total_visa_stoppage_time_calculated': 'Total Visa Wait Time (s)',
            'total_error_wait_time_calculated': 'Total Error Wait Time (s)',
            'booked_by': 'Booked By'
        }
        
        # Write data to worksheet
        for r in dataframe_to_rows(trips_df, index=False, header=True):
            trips_ws.append(r)
        
        # If no data was written, manually add headers
        if trips_ws.max_row == 0:
            trips_ws.append([column_headers.get(col, col) for col in available_columns])
        else:
            # Replace the first row (headers) with user-friendly names
            for col_idx, col_name in enumerate(available_columns, 1):
                if col_name in column_headers:
                    trips_ws.cell(row=1, column=col_idx, value=column_headers[col_name])
        
        # Format Trips sheet
        format_excel_sheet(trips_ws, trips_df)
        
        # Define expected columns for trip legs in the desired order
        expected_leg_columns = [
            'fleet_name', 'trip_id', 'booking_id', 'trip_leg_id', 'trip_leg', 
            'start_time', 'start_time_only', 'end_time', 'end_time_only', 
            'actual_trip_time', 'time_elapsed_obstacle_stoppages', 'time_elapsed_visa_stoppages', 'leg_dispatch_wait_time'
        ]
        
        # Create Trip Legs sheet
        trip_legs_df = pd.DataFrame(trip_legs_data)
        
        # Debug: Log what columns we have
        logging.info(f"Trip legs DataFrame columns: {list(trip_legs_df.columns)}")
        logging.info(f"Expected columns: {expected_leg_columns}")
        
        # Always ensure headers are present, even with empty data
        if trip_legs_df.empty:
            # Create empty DataFrame with expected columns for trip legs
            trip_legs_df = pd.DataFrame(columns=expected_leg_columns)
        else:
            # Reorder DataFrame columns to match the expected order
            trip_legs_df = trip_legs_df[expected_leg_columns]
            
            # Sort by trip_id and trip_leg_id to keep legs of same trip together
            trip_legs_df = trip_legs_df.sort_values(['trip_id', 'trip_leg_id'])
        
        trip_legs_ws = wb.create_sheet("Trip Legs")
        
        # Define user-friendly column headers for Trip Legs
        leg_column_headers = [
            'Fleet', 'Trip ID', 'Booking ID', 'Trip Leg ID', 'Trip Leg', 
            'Leg Start Date', 'Leg Start Time', 'Leg End Date', 'Leg End Time', 
            'Trip Leg Time (s)', 'Obstacle Stoppage Time (s)', 'Visa Wait Time (s)', 'Leg Dispatch Wait Time (s)'
        ]
        
        # Write headers first
        for col_idx, header in enumerate(leg_column_headers, 1):
            trip_legs_ws.cell(row=1, column=col_idx, value=header)
        
        # Write data rows
        row_num = 2
        for leg_data in trip_legs_data:
            trip_legs_ws.cell(row=row_num, column=1, value=leg_data.get('fleet_name'))
            trip_legs_ws.cell(row=row_num, column=2, value=leg_data.get('trip_id'))
            trip_legs_ws.cell(row=row_num, column=3, value=leg_data.get('booking_id'))
            trip_legs_ws.cell(row=row_num, column=4, value=leg_data.get('trip_leg_id'))
            trip_legs_ws.cell(row=row_num, column=5, value=leg_data.get('trip_leg'))
            trip_legs_ws.cell(row=row_num, column=6, value=leg_data.get('start_time'))  # Start Date
            trip_legs_ws.cell(row=row_num, column=7, value=leg_data.get('start_time_only'))  # Start Time
            trip_legs_ws.cell(row=row_num, column=8, value=leg_data.get('end_time'))  # End Date
            trip_legs_ws.cell(row=row_num, column=9, value=leg_data.get('end_time_only'))  # End Time
            trip_legs_ws.cell(row=row_num, column=10, value=leg_data.get('actual_trip_time'))
            trip_legs_ws.cell(row=row_num, column=11, value=leg_data.get('time_elapsed_obstacle_stoppages'))
            trip_legs_ws.cell(row=row_num, column=12, value=leg_data.get('time_elapsed_visa_stoppages'))
            trip_legs_ws.cell(row=row_num, column=13, value=leg_data.get('leg_dispatch_wait_time'))
            row_num += 1
        
        # Format Trip Legs sheet
        format_excel_sheet(trip_legs_ws, trip_legs_df)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
    output.seek(0)
        
    return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=analytics_export.xlsx"},
    )



@router.post("/pause_resume_schedule_trip")
async def pause_schedule_trip(
    pause_schedule_trip_req: rqm.PauseResumeScheduleTripReq,
    user=Depends(AuthValidator('manage_schedule')),
):
    response = {}
    if not user:
        dpd.raise_error("Unknown requester", 401)
    user_name = user["user_name"]

    with DBSession(engine=ccm.engine) as dbsession:
        trips : tm.Trip = dbsession.get_trips_with_booking_id(pause_schedule_trip_req.booking_id)

        if trips is None:
            dpd.raise_error(f"Trip with booking id:{pause_schedule_trip_req.booking_id} does not exist")
        
        trip : tm.Trip = trips[-1]

        if trip.scheduled is False:
            dpd.raise_error(
                f"Trip with booking id:{pause_schedule_trip_req.booking_id} is not scheduled"
            )

        pause_trip = dbsession.get_paused_trip_with_booking_id(pause_schedule_trip_req.booking_id)
        
        if pause_schedule_trip_req.pause:
            if pause_trip is not None:
                dpd.raise_error(f"Trip with booking id:{pause_schedule_trip_req.booking_id} is already paused")
            
            new_trip_metadata = trip.trip_metadata
            if new_trip_metadata.get("total_trip_progress"):
                del new_trip_metadata["total_trip_progress"]
            dbsession.create_paused_trip(
                trip.route,
                trip.priority,
                new_trip_metadata,
                trip.booking_id,
                trip.fleet_name,
                trip.booked_by,
                trip.booking_time,
            )
            for t in trips:
                if t.status in tm.YET_TO_START_TRIP_STATUS:
                    delete_booked_trip_req: rqm.DeleteBookedTripReq = rqm.DeleteBookedTripReq(
                    booking_id=pause_schedule_trip_req.booking_id, trip_id=t.id)
                    response = await dpd.process_req_with_response(None, delete_booked_trip_req, user_name)
                if t.status in tm.ONGOING_TRIP_STATUS:
                    t.trip_metadata["scheduled_end_time"] = trip.trip_metadata["scheduled_start_time"]
                    flag_modified(t, "trip_metadata")
        else:
            if pause_trip is None:
                dpd.raise_error(f"Trip with booking id:{pause_schedule_trip_req.booking_id} is not paused")

            new_trip_metadata = pause_trip.trip_metadata
            new_trip_metadata = tu.modify_trip_metadata(new_trip_metadata)

            new_trip: tm.Trip = dbsession.create_trip(
                pause_trip.route,
                pause_trip.priority,
                new_trip_metadata,
                pause_trip.booking_id,
                pause_trip.fleet_name,
                pause_trip.booked_by,
            )
            dbsession.create_pending_trip(new_trip.id)
            dbsession.delete_paused_trip(pause_trip)

    return response


@router.get("/get_curr_station_of_ongoing_trip_which_is_waiting_for_dispatch/{station_name}")
async def get_curr_station_of_ongoing_trip(
    station_name: str,
    user=Depends(AuthValidator('fm')),
):
    response = {
        "entity_name": None,
        "status": False,
    }
    if not user:
        dpd.raise_error("Unknown requester", 401)

    with DBSession(engine=ccm.engine) as dbsession:
        ongoing_trips: list[tm.OngoingTrip] = dbsession.get_ongoing_trip_with_waiting_station_dispatch_start()
        
        for ongoing_trip in ongoing_trips:
            curr_station = ongoing_trip.curr_station()
            if curr_station == station_name:
                response = {
                    "entity_name": ongoing_trip.sherpa_name,
                    "status": True,
                }
                break

    return response